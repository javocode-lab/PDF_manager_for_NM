from __future__ import annotations

import argparse
import csv
import io
import os
import re
import shutil
import unicodedata
import zipfile
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple, List, Any

from pypdf import PdfReader

PacienteData = Tuple[str, str]
PrefixResult = Tuple[str, str]

DEFAULT_SPECIALTY_PREFIXES: Dict[str, str] = {
    # Especialidades tal como aparecen en los PDF o en prof.txt.
    # La comparación se normaliza, por eso soporta acentos raros como "Clìnicos".
    'FONOAUDIOLOGIA': 'FONO',
    'FONOAUDIOLOGÍA': 'FONO',
    'FONO': 'FONO',
    # Kinesiología genérica NO se infiere automáticamente.
    # KTM/KTR/NEUROKINE deben venir de la tabla maestra de profesionales.
    'KTR': 'KTR',
    'KINESIOTERAPIA RESPIRATORIA': 'KTR',
    'KINESIOLOGIA RESPIRATORIA': 'KTR',
    'KTM': 'KTM',
    'KINESIOTERAPIA MOTORA': 'KTM',
    'KINESIOLOGIA MOTORA': 'KTM',
    'NEUROKINE': 'NEUROKINE',
    'NEUROREHAB': 'NEUROREHAB',
    'NEUROREHABILITACION': 'NEUROREHAB',
    'ENF PROFESIONAL': 'ENF',
    'ENFERMERIA PROFESIONAL': 'ENF',
    'ENFERMERÍA PROFESIONAL': 'ENF',
    'ENFERMERIA': 'ENF',
    'ENFERMERÍA': 'ENF',
    'ENF': 'ENF',
    'CUIDADORES DOMICILIARIOS': 'CUID',
    'CUIDADOR DOMICILIARIO': 'CUID',
    'CUIDADOS DOMICILIARIOS': 'CUID',
    'CUID': 'CUID',
    'MEDICOS CLINICOS': 'MED',
    'MÉDICOS CLÍNICOS': 'MED',
    'MEDICOS CLÍNICOS': 'MED',
    'MÉDICOS CLINICOS': 'MED',
    'MEDICO CLINICO': 'MED',
    'MEDICA CLINICA': 'MED',
    'MEDICINA CLINICA': 'MED',
    'MEDICINA CLÍNICA': 'MED',
    'MEDICINA': 'MED',
    'MEDICO': 'MED',
    'MÉDICO': 'MED',
    'MED': 'MED',
    'MED PED': 'MED_PED',
    'MED_PED': 'MED_PED',
    'PEDIATRIA': 'MED_PED',
    'PEDIATRÍA': 'MED_PED',
    'PSICOLOGIA': 'PSICO',
    'PSICOLOGÍA': 'PSICO',
    'PSICO': 'PSICO',
    'PSICOMOTRICIDAD': 'PSICOMOT',
    'PSICOMOT': 'PSICOMOT',
    'TERAPIA OCUPACIONAL': 'TO',
    'TO': 'TO',
    'NUTRICION': 'NUTRI',
    'NUTRICIÓN': 'NUTRI',
    'NUTRIOLOGIA': 'NUTRI',
    'NUTRIOLOGÍA': 'NUTRI',
    'NUTRI': 'NUTRI',
    'ESTIMULACION TEMPRANA': 'ET',
    'ESTIMULACIÓN TEMPRANA': 'ET',
    'ET': 'ET',
    'NEUROPSICOLOGIA': 'NEUROPSICO',
    'NEUROPSICOLOGÍA': 'NEUROPSICO',
    'NEUROPSICO': 'NEUROPSICO',
    'NEUROLOGIA': 'NEURO',
    'NEUROLOGÍA': 'NEURO',
    'NEURO': 'NEURO',
}

VERSION = 'v0.6.0'

# Especialidades genéricas que no alcanzan para decidir prefijo.
# Ejemplo: un PDF puede decir solo 'Kinesiología', pero el prefijo real puede ser KTM, KTR o NEUROKINE.
AMBIGUOUS_SPECIALTY_KEYS = {
    'KINESIOLOGIA',
    'KINESIOLOGÍA',
    'KINESIOLOGIA GENERAL',
    'KINE',
    'KINESIOLOGOS',
    'KINESIÓLOGOS',
}

ACCENT_TRANSLATION = str.maketrans({
    'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U', 'Ñ': 'N',
    'á': 'A', 'é': 'E', 'í': 'I', 'ó': 'O', 'ú': 'U', 'ü': 'U', 'ñ': 'N',
})

NOISE_PROFESSIONAL_PREFIXES = [
    r'^USUARIO\s+',
    r'^USUSARIO\s+',
    r'^[A-Z]{1,4}\d*\s+MORACHI,\s*',
    r'^ALVARADO\d*[-\s]*(TM|TT|TN)?[,]?\s*',
]


def normalize_spaces(value: str) -> str:
    value = value or ''
    value = value.replace('﻿', '').replace('​', '').replace(' ', ' ')
    return re.sub(r'\s+', ' ', value).strip()


def remove_accents_upper(value: str) -> str:
    # Normalización Unicode real: soporta á, ñ, ü y también acentos raros como ì.
    normalized = unicodedata.normalize('NFD', value or '')
    without_marks = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    return without_marks.upper()


def normalize_key(value: str) -> str:
    value = remove_accents_upper(normalize_spaces(value))
    value = value.replace('.', '')
    value = re.sub(r'\s*,\s*', ', ', value)
    value = re.sub(r'\s*/\s*', '/', value)
    value = re.sub(r'[^A-Z0-9,/_ -]+', ' ', value)
    return normalize_spaces(value)


def compact_key(value: str) -> str:
    return re.sub(r'[^A-Z0-9]+', '', normalize_key(value))


def specialty_key_variants(value: str) -> set[str]:
    base = normalize_key(value)
    compact = compact_key(base)
    variants = {base, compact} if base else set()

    # Equivalencias frecuentes que aparecen con textos levemente distintos.
    if 'MEDICOS' in base and 'CLINICOS' in base:
        variants.update({'MEDICOS CLINICOS', 'MEDICOSCLINICOS'})
    if 'ENF' in base or 'ENFERMERIA' in base:
        variants.update({'ENF PROFESIONAL', 'ENFPROFESIONAL', 'ENFERMERIA', 'ENFERMERIA PROFESIONAL'})
    if 'CUIDADOR' in base or 'CUIDADORES' in base or 'CUIDADOS' in base:
        variants.update({'CUIDADORES DOMICILIARIOS', 'CUIDADORESDOMICILIARIOS', 'CUIDADOS DOMICILIARIOS'})
    if 'FONOAUDIOLOG' in base:
        variants.update({'FONOAUDIOLOGIA', 'FONOAUDIOLOGIA'})
    if 'NUTRI' in base:
        variants.update({'NUTRICION', 'NUTRIOLOGIA', 'NUTRI'})
    if 'PSICOMOT' in base:
        variants.update({'PSICOMOTRICIDAD', 'PSICOMOT'})
    if 'PSICOLOG' in base:
        variants.update({'PSICOLOGIA', 'PSICO'})
    if 'TERAPIA' in base and 'OCUPACIONAL' in base:
        variants.update({'TERAPIA OCUPACIONAL', 'TO'})
    if 'KINESIO' in base:
        variants.update({'KINESIOLOGIA', 'KINE'})

    return {v for v in variants if v}


def professional_key_variants(value: str) -> set[str]:
    base = normalize_key(value)
    variants = {base, compact_key(base)} if base else set()

    cleaned = base
    for pattern in NOISE_PROFESSIONAL_PREFIXES:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE).strip()
    if cleaned and cleaned != base:
        variants.add(cleaned)
        variants.add(compact_key(cleaned))

    # Algunos PDFs pueden traer dobles comas por prefijos operativos.
    if ',' in base:
        parts = [normalize_spaces(p) for p in base.split(',') if normalize_spaces(p)]
        if len(parts) >= 2:
            last_two = ', '.join(parts[-2:])
            variants.add(last_two)
            variants.add(compact_key(last_two))

    return {v for v in variants if v}


def sanitize_name(value: str) -> str:
    value = normalize_spaces(value)
    value = re.sub(r'[<>:"/\\|?*\x00-\x1F]', '-', value)
    value = value.strip(' .')
    return value or 'SIN_NOMBRE'


def sanitize_prefix(value: str) -> str:
    value = remove_accents_upper(normalize_spaces(value))
    value = re.sub(r'[^A-Z0-9]+', '_', value)
    value = value.strip('_')
    return value or 'SIN_PREFIJO'


def split_mapping_line(line: str) -> list[str]:
    raw = (line or '').replace('﻿', '').replace('​', '').strip()
    if not raw or raw.startswith('#'):
        return []

    # Primero respeta tabulaciones reales. No normalizamos antes porque eso las borra.
    if '	' in raw:
        parts = [normalize_spaces(p) for p in raw.split('	')]
        parts = [p for p in parts if p and p != '-']
        if len(parts) >= 2:
            return parts

    line = normalize_spaces(raw)

    for separator in [';', '|', '=']:
        if separator in line:
            parts = [normalize_spaces(p) for p in line.split(separator) if normalize_spaces(p)]
            if len(parts) >= 2:
                return parts

    # Formato típico: PROFESIONAL - SIGLA o CODIGO - PACIENTE - ENTIDAD.
    return [normalize_spaces(p) for p in re.split(r'\s+-\s+', line) if normalize_spaces(p)]


def parse_data_txt(path: Optional[Path]) -> Dict[str, PacienteData]:
    mapping: Dict[str, PacienteData] = {}
    if not path or not path.exists():
        return mapping

    text = path.read_text(encoding='utf-8', errors='replace')
    for raw in text.splitlines():
        line = normalize_spaces(raw)
        if not line or not re.match(r'^\d+', line):
            continue

        parts = split_mapping_line(line)
        if len(parts) >= 3 and re.fullmatch(r'\d+', parts[0]):
            codigo, nombre, entidad = parts[0], parts[1], parts[2]
            mapping[codigo] = (sanitize_name(nombre), sanitize_name(entidad))
            continue

        m = re.match(r'^(\d+)\s*-\s*(.*?)\s*-\s*(.+?)\s*$', line)
        if m:
            codigo, nombre, entidad = m.groups()
            mapping[codigo] = (sanitize_name(nombre), sanitize_name(entidad))

    return mapping


def parse_prof_txt(path: Optional[Path]) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    """Devuelve mapas por profesional, matrícula y especialidad/prefijo."""
    professional_map: Dict[str, str] = {}
    license_map: Dict[str, str] = {}
    specialty_map: Dict[str, str] = {}

    if not path or not path.exists():
        return professional_map, license_map, specialty_map

    text = path.read_text(encoding='utf-8', errors='replace')
    for raw in text.splitlines():
        line = normalize_spaces(raw)
        if not line or line.startswith('#'):
            continue

        lower = remove_accents_upper(line)
        if lower.startswith('PROFESIONAL') or lower.startswith('NOMBRE') or lower.startswith('MATRICULA'):
            continue

        parts = split_mapping_line(line)
        if len(parts) < 2:
            continue

        prefix = sanitize_prefix(parts[-1])

        if len(parts) == 2:
            key = parts[0]
            for variant in professional_key_variants(key):
                professional_map[variant] = prefix
            specialty_map[normalize_key(key)] = prefix
            specialty_map[compact_key(key)] = prefix
            continue

        # Formato alternativo: matricula - profesional - prefijo.
        first_key = parts[0]
        middle_key = parts[-2]
        if re.fullmatch(r'[A-Z]?\d[\d/.-]*[A-Z]?', first_key.strip(), flags=re.IGNORECASE):
            license_map[normalize_key(first_key)] = prefix
            license_map[compact_key(first_key)] = prefix

        for variant in professional_key_variants(middle_key):
            professional_map[variant] = prefix

    return professional_map, license_map, specialty_map


def read_xlsx_rows(path: Path) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            'Para leer archivos .xlsx instalá las dependencias con: py -m pip install -r requirements.txt'
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[List[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        if any(v is not None and str(v).strip() for v in values):
            rows.append(values)
    workbook.close()
    return rows


def cell_to_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_spaces(str(value))


def parse_patient_xlsx(path: Optional[Path]) -> Dict[str, PacienteData]:
    mapping: Dict[str, PacienteData] = {}
    if not path or not path.exists():
        return mapping

    rows = read_xlsx_rows(path)
    if not rows:
        return mapping

    headers = [normalize_key(cell_to_text(v)) for v in rows[0]]

    def find_col(possible_names: List[str], fallback: int) -> int:
        for name in possible_names:
            key = normalize_key(name)
            for idx, header in enumerate(headers):
                if key in header or header in key:
                    return idx
        return fallback

    code_col = find_col(['CODIGO', 'CODIGO PACIENTE', 'CODIGO_PACIENTE'], 0)
    name_col = find_col(['NOMBRE', 'PACIENTE'], 2)
    entity_col = find_col(['ENTIDAD', 'OBRA SOCIAL'], 4)

    for row in rows[1:]:
        code = cell_to_text(row[code_col] if code_col < len(row) else '')
        name = cell_to_text(row[name_col] if name_col < len(row) else '')
        entity = cell_to_text(row[entity_col] if entity_col < len(row) else '')

        code_match = re.search(r'\d+', code)
        if not code_match or not name:
            continue

        mapping[code_match.group(0)] = (sanitize_name(name), sanitize_name(entity or 'SIN_ENTIDAD'))

    return mapping


def parse_prof_xlsx(path: Optional[Path]) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    professional_map: Dict[str, str] = {}
    license_map: Dict[str, str] = {}
    specialty_map: Dict[str, str] = {}

    if not path or not path.exists():
        return professional_map, license_map, specialty_map

    rows = read_xlsx_rows(path)
    if not rows:
        return professional_map, license_map, specialty_map

    headers = [normalize_key(cell_to_text(v)) for v in rows[0]]

    def find_col(possible_names: List[str], fallback: int) -> int:
        for name in possible_names:
            key = normalize_key(name)
            for idx, header in enumerate(headers):
                if key in header or header in key:
                    return idx
        return fallback

    prof_col = find_col(['PROFESIONAL', 'NOMBRE'], 0)
    prefix_col = find_col(['PREFIJO', 'ESPECIALIDAD', 'SIGLA'], 2)
    license_col = find_col(['MATRICULA', 'MATRÍCULA'], -1)

    for row in rows[1:]:
        professional = cell_to_text(row[prof_col] if prof_col >= 0 and prof_col < len(row) else '')
        prefix = cell_to_text(row[prefix_col] if prefix_col >= 0 and prefix_col < len(row) else '')
        license_number = cell_to_text(row[license_col] if license_col >= 0 and license_col < len(row) else '')

        if not professional or not prefix:
            continue

        sanitized_prefix = sanitize_prefix(prefix)
        for variant in professional_key_variants(professional):
            professional_map[variant] = sanitized_prefix

        if license_number:
            license_map[normalize_key(license_number)] = sanitized_prefix
            license_map[compact_key(license_number)] = sanitized_prefix

        # Si el Excel tiene una fila donde la primera columna es una especialidad genérica, también se soporta.
        # Para profesionales reales, esto no afecta porque se resuelve por nombre primero.
        first_key = normalize_key(professional)
        if first_key in DEFAULT_SPECIALTY_PREFIXES or any(word in first_key for word in ['ENFERMERIA', 'FONOAUDIOLOG', 'MEDICOS', 'TERAPIA OCUPACIONAL', 'PSICOLOG', 'NUTRI']):
            specialty_map[first_key] = sanitized_prefix
            specialty_map[compact_key(first_key)] = sanitized_prefix

    return professional_map, license_map, specialty_map


def parse_patient_file(path: Optional[Path]) -> Dict[str, PacienteData]:
    if not path:
        return {}
    suffix = path.suffix.lower()
    if suffix == '.xlsx':
        return parse_patient_xlsx(path)
    return parse_data_txt(path)


def parse_prof_file(path: Optional[Path]) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    if not path:
        return {}, {}, {}
    suffix = path.suffix.lower()
    if suffix == '.xlsx':
        return parse_prof_xlsx(path)
    return parse_prof_txt(path)


def extract_code_from_filename(filename: str) -> Optional[str]:
    base = os.path.basename(filename)
    m = re.search(r'Paciente[_\s-]*(\d+)', base, flags=re.IGNORECASE)
    return m.group(1) if m else None


def extract_pdf_text_from_bytes(pdf_bytes: bytes, max_pages: int = 1) -> str:
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        return '\n'.join(page.extract_text() or '' for page in reader.pages[:max_pages])
    except Exception:
        return ''


def extract_header_slice(pdf_text: str) -> str:
    text = pdf_text[:5000]
    m = re.search(r'(.*?Periodo:.*?Especialidad:.*?)(?:Fecha\s+Visita|Motivo\s+Internacion|Diagnostico\s+Ingreso)', text, flags=re.IGNORECASE | re.DOTALL)
    if m:
        return m.group(1)

    m = re.search(r'(.*?Fecha\s+Visita)', text, flags=re.IGNORECASE | re.DOTALL)
    if m:
        return m.group(1)

    return text


def infer_patient_from_pdf_text(code: Optional[str], pdf_text: str) -> Optional[PacienteData]:
    if not pdf_text:
        return None

    nombre = None
    entidad = None

    if code:
        m_name = re.search(rf'\b{re.escape(code)}\b\s*-\s*([^\n\r]+)', pdf_text, flags=re.IGNORECASE)
        if m_name:
            nombre = sanitize_name(m_name.group(1))

    # Patrón típico: Paciente: Obra Social: MEDIFE CBA Nro. Afiliado
    m_entity = re.search(r'Obra\s+Social:\s*(.*?)\s+Nro\.?\s*Afiliado', pdf_text, flags=re.IGNORECASE | re.DOTALL)
    if m_entity:
        entidad = sanitize_name(m_entity.group(1))

    if nombre and entidad:
        return nombre, entidad

    return None


def extract_professional_data(pdf_text: str) -> Tuple[str, str, str]:
    """Devuelve: profesional, especialidad, matrícula."""
    header = extract_header_slice(pdf_text)
    professional = ''
    specialty = ''
    license_number = ''

    m_license = re.search(r'Matricula:\s*([A-Z]?\d[\d/.-]*[A-Z]?)', header, flags=re.IGNORECASE)
    if m_license:
        license_number = normalize_spaces(m_license.group(1))

    m_specialty = re.search(
        r'Especialidad:\s*(.*?)(?:\n|MNA\.|Fecha\s+Visita|Motivo\s+Internacion|Diagnostico\s+Ingreso|$)',
        header,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if m_specialty:
        specialty = sanitize_name(m_specialty.group(1))

    # Patrón robusto para DNI. Acepta DNI numérico y variantes como "DNI a35674058".
    for raw_line in header.splitlines():
        line = normalize_spaces(raw_line)
        if re.search(r'\bDNI\b\s*[A-Z]?\d{6,9}\b', line, flags=re.IGNORECASE):
            candidate = re.sub(r'\bDNI\b\s*[A-Z]?\d{6,9}\b.*$', '', line, flags=re.IGNORECASE)
            candidate = sanitize_name(candidate)
            if candidate and len(candidate) >= 4:
                professional = candidate
                break

    # Fallback más amplio: si la línea contiene DNI pero el documento tiene algún carácter raro.
    if not professional:
        for raw_line in header.splitlines():
            line = normalize_spaces(raw_line)
            if re.search(r'\bDNI\b', line, flags=re.IGNORECASE):
                candidate = re.sub(r'\bDNI\b.*$', '', line, flags=re.IGNORECASE)
                candidate = sanitize_name(candidate)
                if candidate and len(candidate) >= 4:
                    professional = candidate
                    break

    return professional, specialty, license_number


def find_prefix(
    professional: str,
    specialty: str,
    license_number: str,
    professional_map: Dict[str, str],
    license_map: Dict[str, str],
    specialty_map: Dict[str, str],
) -> PrefixResult:
    if license_number:
        for key in {normalize_key(license_number), compact_key(license_number)}:
            if key in license_map:
                return license_map[key], 'prof_txt_matricula'

    if professional:
        for key in professional_key_variants(professional):
            if key in professional_map:
                return professional_map[key], 'prof_txt_profesional'

        # Fallback fuzzy controlado: solo si la similitud es muy alta.
        candidate_compact = compact_key(professional)
        best_key = ''
        best_ratio = 0.0
        for map_key in professional_map:
            map_compact = compact_key(map_key)
            if not candidate_compact or not map_compact:
                continue
            ratio = SequenceMatcher(None, candidate_compact, map_compact).ratio()
            contains = candidate_compact in map_compact or map_compact in candidate_compact
            if (ratio > best_ratio) and (ratio >= 0.94 or (contains and min(len(candidate_compact), len(map_compact)) >= 10)):
                best_ratio = ratio
                best_key = map_key
        if best_key:
            return professional_map[best_key], f'prof_txt_profesional_fuzzy_{best_ratio:.2f}'

    if specialty:
        specialty_keys = specialty_key_variants(specialty)
        ambiguous_keys = {normalize_key(k) for k in AMBIGUOUS_SPECIALTY_KEYS} | {compact_key(k) for k in AMBIGUOUS_SPECIALTY_KEYS}

        # Primero se respetan especialidades explícitas cargadas en la tabla maestra.
        for key in specialty_keys:
            if key in specialty_map:
                return specialty_map[key], 'tabla_profesionales_especialidad'

        # Si la especialidad es genérica de kinesiología, NO se infiere el subtipo.
        if any(key in ambiguous_keys for key in specialty_keys):
            return 'SIN_PREFIJO', 'especialidad_kinesiologia_ambigua_requiere_tabla_profesionales'

        # Fallback automático solo para especialidades no ambiguas.
        for key in specialty_keys:
            if key in DEFAULT_SPECIALTY_PREFIXES:
                return DEFAULT_SPECIALTY_PREFIXES[key], 'prefijo_default_especialidad_no_ambigua'

        compact_specialty = compact_key(specialty)
        for default_key, default_prefix in DEFAULT_SPECIALTY_PREFIXES.items():
            compact_default = compact_key(default_key)
            if compact_default and (compact_default in compact_specialty or compact_specialty in compact_default):
                # No inferir kinesiología si no hay subtipo explícito.
                if default_prefix in {'KTR', 'KTM', 'NEUROKINE', 'NEUROREHAB'} and 'RESP' not in compact_specialty and 'MOTOR' not in compact_specialty and 'NEURO' not in compact_specialty:
                    continue
                return default_prefix, 'prefijo_default_especialidad_contiene_no_ambigua'

    return 'SIN_PREFIJO', 'sin_coincidencia_en_tabla_profesionales'


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem, suffix = path.stem, path.suffix
    counter = 2
    while True:
        candidate = path.with_name(f'{stem} ({counter}){suffix}')
        if not candidate.exists():
            return candidate
        counter += 1


def zip_dir(source_dir: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in sorted(source_dir.rglob('*')):
            if file.is_file():
                zf.write(file, file.relative_to(source_dir).as_posix())


def iter_input_pdfs_from_zip(input_zip: Path) -> Iterable[Tuple[str, bytes]]:
    with zipfile.ZipFile(input_zip, 'r') as zin:
        for info in zin.infolist():
            if info.is_dir():
                continue
            original_name = os.path.basename(info.filename)
            if original_name.lower().endswith('.pdf'):
                yield original_name, zin.read(info)


def strip_previous_prefixes(filename: str) -> str:
    """Evita nombres tipo FONO - ALMADA - HC_... si se reprocesa una salida vieja."""
    base = os.path.basename(filename)
    prefixes = ['SIN_PREFIJO'] + sorted(set(DEFAULT_SPECIALTY_PREFIXES.values()), key=len, reverse=True)
    escaped = '|'.join(re.escape(p) for p in prefixes)
    pattern = rf'^({escaped})\s+-\s+'
    while re.match(pattern, base, flags=re.IGNORECASE):
        base = re.sub(pattern, '', base, count=1, flags=re.IGNORECASE)

    # También elimina prefijos de apellido generados por la versión anterior: "ALMADA - HC_...".
    base = re.sub(r'^[A-ZÁÉÍÓÚÜÑ][A-ZÁÉÍÓÚÜÑ ]{2,}\s+-\s+(HC_Profesional_)', r'\1', base, count=1)
    return base


def organize(
    input_zip: Path,
    output_dir: Path,
    data_txt: Optional[Path] = None,
    prof_txt: Optional[Path] = None,
    output_zip: Optional[Path] = None,
) -> None:
    if not input_zip.exists():
        raise FileNotFoundError(f'No se encontró el ZIP de entrada: {input_zip}')

    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True)

    patient_mapping = parse_patient_file(data_txt)
    professional_map, license_map, specialty_map = parse_prof_file(prof_txt)

    registros = []
    prefix_not_found = []
    debug_headers_sin_prefijo = []
    patient_from_pdf = 0
    patient_from_txt = 0
    patient_not_found = []
    entity_counter = Counter()
    prefix_counter = Counter()
    source_prefix_counter = Counter()
    total_files = 0

    for original_name_from_zip, pdf_bytes in iter_input_pdfs_from_zip(input_zip):
        total_files += 1
        original_name = strip_previous_prefixes(original_name_from_zip)
        code = extract_code_from_filename(original_name)
        pdf_text = extract_pdf_text_from_bytes(pdf_bytes)

        inferred_data = infer_patient_from_pdf_text(code, pdf_text)
        if inferred_data:
            nombre, entidad = inferred_data
            patient_source = 'pdf_header'
            patient_from_pdf += 1
        elif code and code in patient_mapping:
            nombre, entidad = patient_mapping[code]
            patient_source = 'tabla_pacientes_fallback'
            patient_from_txt += 1
        else:
            nombre = f'PACIENTE_{code}' if code else 'SIN_CODIGO'
            entidad = 'SIN_CLASIFICAR'
            patient_source = 'sin_coincidencia'
            patient_not_found.append(original_name_from_zip)

        professional, specialty, license_number = extract_professional_data(pdf_text)
        prefix, prefix_source = find_prefix(
            professional,
            specialty,
            license_number,
            professional_map,
            license_map,
            specialty_map,
        )
        if prefix == 'SIN_PREFIJO':
            prefix_not_found.append(original_name_from_zip)
            debug_headers_sin_prefijo.append((original_name_from_zip, extract_header_slice(pdf_text)[:1800]))

        final_filename = sanitize_name(f'{prefix} - {Path(original_name).stem}') + Path(original_name).suffix.lower()
        final_dir = output_dir / sanitize_name(entidad) / sanitize_name(nombre)
        final_dir.mkdir(parents=True, exist_ok=True)
        final_path = unique_path(final_dir / final_filename)
        final_path.write_bytes(pdf_bytes)

        entity_counter[entidad] += 1
        prefix_counter[prefix] += 1
        source_prefix_counter[prefix_source] += 1

        registros.append({
            'codigo_paciente': code or '',
            'entidad': entidad,
            'paciente': nombre,
            'profesional_detectado': professional,
            'especialidad_detectada': specialty,
            'matricula_detectada': license_number,
            'prefijo_asignado': prefix,
            'archivo_original': original_name_from_zip,
            'archivo_base_limpio': original_name,
            'archivo_final': final_path.name,
            'ruta_final': final_path.relative_to(output_dir).as_posix(),
            'fuente_paciente': patient_source,
            'fuente_prefijo': prefix_source,
        })

    fieldnames = [
        'codigo_paciente', 'entidad', 'paciente', 'profesional_detectado',
        'especialidad_detectada', 'matricula_detectada', 'prefijo_asignado',
        'archivo_original', 'archivo_base_limpio', 'archivo_final', 'ruta_final',
        'fuente_paciente', 'fuente_prefijo',
    ]

    with (output_dir / 'reporte_organizacion.csv').open('w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(registros)

    professionals_detected = sorted({
        (r['profesional_detectado'], r['especialidad_detectada'], r['matricula_detectada'], r['prefijo_asignado'], r['fuente_prefijo'])
        for r in registros
        if r['profesional_detectado'] or r['especialidad_detectada'] or r['matricula_detectada']
    })

    with (output_dir / 'profesionales_detectados.csv').open('w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['profesional_detectado', 'especialidad_detectada', 'matricula_detectada', 'prefijo_asignado', 'fuente_prefijo', 'linea_sugerida_prof_txt'])
        for professional, specialty, license_number, prefix, source in professionals_detected:
            suggested_key = professional or specialty or license_number
            suggested_prefix = prefix if prefix != 'SIN_PREFIJO' else 'COMPLETAR'
            writer.writerow([professional, specialty, license_number, prefix, source, f'{suggested_key} - {suggested_prefix}'])

    with (output_dir / 'archivos_sin_prefijo.csv').open('w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows([r for r in registros if r['prefijo_asignado'] == 'SIN_PREFIJO'])

    if debug_headers_sin_prefijo:
        with (output_dir / 'debug_headers_sin_prefijo.txt').open('w', encoding='utf-8') as f:
            for filename, header in debug_headers_sin_prefijo:
                f.write('=' * 90 + '\n')
                f.write(f'ARCHIVO: {filename}\n')
                f.write('-' * 90 + '\n')
                f.write(header.strip() + '\n\n')

    with (output_dir / 'reporte_organizacion.txt').open('w', encoding='utf-8') as f:
        f.write('REPORTE DE ORGANIZACIÓN DE PDFs POR ENTIDAD, PACIENTE Y PREFIJO PROFESIONAL\n')
        f.write('=' * 82 + '\n\n')
        f.write(f'PDFs encontrados: {total_files}\n')
        f.write(f'PDFs organizados: {len(registros)}\n')
        f.write(f'PDFs sin clasificar por paciente: {len(patient_not_found)}\n')
        f.write(f'PDFs sin prefijo profesional: {len(prefix_not_found)}\n')
        f.write(f'Pacientes obtenidos desde PDF: {patient_from_pdf}\n')
        f.write(f'Pacientes obtenidos desde datos.txt fallback: {patient_from_txt}\n')
        f.write(f'Profesionales cargados desde tabla profesional: {len(professional_map)} claves internas\n')
        f.write(f'Especialidades default disponibles: {len(DEFAULT_SPECIALTY_PREFIXES)}\n')
        f.write(f'Versión del organizador: {VERSION}\n\n')

        f.write('DISTRIBUCIÓN POR ENTIDAD\n')
        f.write('-' * 25 + '\n')
        for entidad, count in sorted(entity_counter.items()):
            f.write(f'{entidad}: {count}\n')

        f.write('\nDISTRIBUCIÓN POR PREFIJO\n')
        f.write('-' * 25 + '\n')
        for prefix, count in sorted(prefix_counter.items()):
            f.write(f'{prefix}: {count}\n')

        f.write('\nFUENTES DE PREFIJO\n')
        f.write('-' * 25 + '\n')
        for source, count in sorted(source_prefix_counter.items()):
            f.write(f'{source}: {count}\n')

        if prefix_not_found:
            f.write('\nARCHIVOS SIN PREFIJO PROFESIONAL\n')
            f.write('-' * 35 + '\n')
            for name in prefix_not_found:
                f.write(f'{name}\n')
            f.write('\nVer detalle en: archivos_sin_prefijo.csv\n')

        if patient_not_found:
            f.write('\nARCHIVOS SIN CLASIFICAR POR PACIENTE\n')
            f.write('-' * 38 + '\n')
            for name in patient_not_found:
                f.write(f'{name}\n')

    if output_zip:
        zip_dir(output_dir, output_zip)
        print(f'Organización finalizada. Carpeta generada: {output_dir}')
        print(f'ZIP generado: {output_zip}')
    else:
        print(f'Organización finalizada. Carpeta generada: {output_dir}')
        print('No se generó ZIP porque no se indicó --output-zip.')


def main() -> None:
    parser = argparse.ArgumentParser(
        description=f'Organiza PDFs por entidad/paciente y renombra cada PDF con prefijo profesional. Versión {VERSION}.'
    )
    parser.add_argument('--input-zip', required=True, help='ZIP con PDFs. Puede ser bruto u organizado previamente.')
    parser.add_argument('--patients-file', required=False, default=None, help='Opcional. Excel/TXT con pacientes: codigo, nombre y entidad. Recomendado: Listado_Pacientes.xlsx.')
    parser.add_argument('--professionals-file', required=False, default=None, help='Opcional. Excel/TXT con profesionales y prefijo. Recomendado: Listado_profesionales.xlsx.')
    parser.add_argument('--data-txt', required=False, default=None, help='Alias compatible. TXT/XLSX de pacientes. Si se indica --patients-file, este queda en segundo lugar.')
    parser.add_argument('--prof-txt', required=False, default=None, help='Alias compatible. TXT/XLSX de profesionales. Si se indica --professionals-file, este queda en segundo lugar.')
    parser.add_argument('--output-dir', default='salida_organizada', help='Carpeta final de salida.')
    parser.add_argument('--output-zip', default=None, help='Opcional. Si se indica, también genera ZIP final.')
    args = parser.parse_args()

    organize(
        input_zip=Path(args.input_zip),
        output_dir=Path(args.output_dir),
        data_txt=Path(args.patients_file or args.data_txt) if (args.patients_file or args.data_txt) else None,
        prof_txt=Path(args.professionals_file or args.prof_txt) if (args.professionals_file or args.prof_txt) else None,
        output_zip=Path(args.output_zip) if args.output_zip else None,
    )


if __name__ == '__main__':
    main()
